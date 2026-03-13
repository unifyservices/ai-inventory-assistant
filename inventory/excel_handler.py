import os
import json
from typing import Dict, List, Tuple
from openpyxl import load_workbook as _load_wb
from collections import Counter


def _find_header_row(rows):
    """Find the best header row - the first row with the most non-empty cells."""
    best_idx = 0
    best_count = 0
    for i, row in enumerate(rows[:10]):  # Check first 10 rows
        non_empty = sum(1 for v in row if v is not None and str(v).strip())
        if non_empty > best_count:
            best_count = non_empty
            best_idx = i
    return best_idx


def load_inventory(filepath: str) -> Tuple[List[str], List[dict]]:
    """Load ALL sheets from Excel file. Each record is tagged with _sheet_name."""
    wb = _load_wb(filepath, read_only=True, data_only=True)
    all_headers = set()
    all_data = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue

        # Find the actual header row (may not be row 0 if there are merged/title rows)
        header_idx = _find_header_row(rows)
        header_row = rows[header_idx]

        # Build headers, skipping truly empty columns
        headers = []
        col_indices = []  # Track which column indices have real headers
        for i, h in enumerate(header_row):
            if h is not None and str(h).strip():
                headers.append(str(h).strip())
                col_indices.append(i)

        if not headers:
            continue

        all_headers.update(headers)

        # Read data rows (everything after the header row)
        for idx, row in enumerate(rows[header_idx + 1:], start=header_idx + 2):
            if all(v is None for v in row):
                continue
            record = {"_row_number": idx, "_sheet_name": sheet_name}
            for col_name, col_i in zip(headers, col_indices):
                val = row[col_i] if col_i < len(row) else None
                record[col_name] = val
            # Skip rows where all real columns are empty
            if all(record.get(h) is None for h in headers):
                continue
            all_data.append(record)

    wb.close()
    final_headers = ["_sheet_name"] + sorted(all_headers)
    return final_headers, all_data


def search_records(data: List[dict], filters: dict) -> List[dict]:
    """Filter records where column values contain the search term (case-insensitive).
    Special key '_any' searches across ALL columns."""
    results = data
    for col, term in filters.items():
        if col in ("_row_number",):
            continue
        term_lower = str(term).lower()
        if col == "_any":
            # Search across ALL columns for this term
            results = [
                r for r in results
                if any(
                    v is not None and term_lower in str(v).lower()
                    for k, v in r.items() if not k.startswith("_")
                )
            ]
        else:
            results = [
                r for r in results
                if col in r and r[col] is not None and term_lower in str(r[col]).lower()
            ]
    return results


def get_summary(headers: List[str], data: List[dict]) -> dict:
    """Return summary statistics of the inventory."""
    summary = {
        "total_records": len(data),
        "columns": headers,
        "column_stats": {},
    }
    for col in headers:
        values = [r.get(col) for r in data if r.get(col) is not None]
        numeric_vals = []
        for v in values:
            try:
                numeric_vals.append(float(v))
            except (ValueError, TypeError):
                pass
        if numeric_vals:
            summary["column_stats"][col] = {
                "type": "numeric",
                "min": min(numeric_vals),
                "max": max(numeric_vals),
                "avg": round(sum(numeric_vals) / len(numeric_vals), 2),
                "non_empty": len(values),
            }
        else:
            top_values = Counter(str(v) for v in values).most_common(5)
            summary["column_stats"][col] = {
                "type": "text",
                "unique_values": len(set(str(v) for v in values)),
                "non_empty": len(values),
                "top_values": dict(top_values),
            }
    return summary


def find_low_stock(data: List[dict], quantity_column: str, threshold: int = 10) -> List[dict]:
    """Find records where quantity is below the threshold."""
    results = []
    for r in data:
        val = r.get(quantity_column)
        if val is None:
            continue
        try:
            if float(val) < threshold:
                results.append(r)
        except (ValueError, TypeError):
            continue
    return results


def find_duplicates(data: List[dict], column: str) -> Dict[str, List[dict]]:
    """Find duplicate entries based on a column. Returns {value: [matching rows]}."""
    groups: Dict[str, List[dict]] = {}
    for r in data:
        val = r.get(column)
        if val is None:
            continue
        key = str(val).strip().lower()
        groups.setdefault(key, []).append(r)
    return {k: v for k, v in groups.items() if len(v) > 1}


def update_record(filepath: str, row_number: int, column: str, new_value, sheet_name: str = None) -> dict:
    """Update a specific cell in the Excel file. Returns old and new values."""
    wb = _load_wb(filepath)
    ws = wb[sheet_name] if sheet_name and sheet_name in wb.sheetnames else wb.active
    headers = [str(cell.value).strip() if cell.value else f"Column_{i}"
               for i, cell in enumerate(ws[1])]
    if column not in headers:
        wb.close()
        return {"error": f"Column '{column}' not found. Available: {headers}"}
    col_idx = headers.index(column) + 1
    old_value = ws.cell(row=row_number, column=col_idx).value
    ws.cell(row=row_number, column=col_idx, value=new_value)
    wb.save(filepath)
    wb.close()
    return {"row": row_number, "column": column, "old_value": old_value, "new_value": new_value}


def add_record(filepath: str, record: dict, sheet_name: str = None) -> dict:
    """Append a new row to the Excel file."""
    wb = _load_wb(filepath)
    ws = wb[sheet_name] if sheet_name and sheet_name in wb.sheetnames else wb.active
    headers = [str(cell.value).strip() if cell.value else f"Column_{i}"
               for i, cell in enumerate(ws[1])]
    new_row = [record.get(h) for h in headers]
    ws.append(new_row)
    new_row_number = ws.max_row
    wb.save(filepath)
    wb.close()
    return {"row_number": new_row_number, "record": record}


def format_records(records: List[dict], max_per_sheet: int = 20) -> str:
    """Format records grouped by sheet, limiting per sheet to save tokens."""
    if not records:
        return "No records found."
    # Group by sheet
    by_sheet: Dict[str, List[dict]] = {}
    for r in records:
        sheet = r.get("_sheet_name", "Unknown")
        by_sheet.setdefault(sheet, []).append(r)

    parts = []
    total_shown = 0
    for sheet, rows in by_sheet.items():
        shown = rows[:max_per_sheet]
        clean = [{k: v for k, v in r.items() if k not in ("_row_number",) and v is not None}
                 for r in shown]
        total_shown += len(clean)
        section = f"Sheet: {sheet} ({len(rows)} matches"
        if len(rows) > max_per_sheet:
            section += f", showing first {max_per_sheet}"
        section += ")\n"
        section += json.dumps(clean, indent=2, default=str)
        parts.append(section)

    return "\n\n".join(parts)
